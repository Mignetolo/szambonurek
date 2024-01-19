import pyshark
import re
from openpyxl import Workbook
import hashlib

def initialize_regex():
    """
    Initialize the dictionary of regular expressions for different match types.

    Returns:
        dict: A dictionary containing regular expressions for different match types.
    """
    regex_dict = {
        'IPv4': re.compile(r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}'),
        'IPv6': re.compile(r'(?:[0-9a-fA-F]{1,4}:){7}[0-9a-fA-F]{1,4}'),
        'Email': re.compile(r'\S+@\S+'),
        'Domain Name': re.compile(r'\b(?:[a-z0-9](?:[a-z0-9-]{0,61}[a-z0-9])?\.)+[a-z]{2,}\b(?![\d.])'),
    }
    return regex_dict

def save_packet_data(sheet, packet):
    """
    Save packet data to a worksheet.

    Args:
        sheet (openpyxl.Worksheet): The worksheet object to save the packet data to.
        packet (pyshark.packet.Packet): The packet object containing the data to be saved.

    Returns:
        None

    Raises:
        None

    """
    packet_info = [packet.number, packet.sniff_timestamp, packet.length, packet.transport_layer]
    sheet.append(packet_info)

def process_packets(cap, regex_dict, workbook):
    """
    Process packets from a capture file and generate a summary of matches.

    Args:
        cap (pyshark.FileCapture): The capture file object.
        regex_dict (dict): A dictionary containing regex patterns for different match types.
        workbook (openpyxl.Workbook): The workbook object to store the packet data and summary.

    Returns:
        openpyxl.Workbook: The updated workbook object with packet data and summary.

    Raises:
        None

    """
    summary_dict = {}
    packet_counter = 1

    summary_sheet = workbook.create_sheet(title='Summary')
    summary_sheet.append(['Match Type', 'Match Value', 'Count', 'Packet Numbers'])

    for packet in cap:
        packet_data = str(packet)
        packet_saved = False

        for regex_name, regex_pattern in regex_dict.items():
            matches = regex_pattern.findall(packet_data)
            if matches:
                if not packet_saved:
                    packet_saved = True
                    sheet_name = f'Packet_{packet_counter}'
                    sheet = workbook.create_sheet(title=sheet_name)
                    sheet.append(['Packet Number', 'Time', 'Length', 'Transport Layer'])
                    save_packet_data(sheet, packet)

                data_sheet = workbook[sheet_name]
                data_sheet.append([f'{regex_name} Match'])
                for match in matches:
                    data_sheet.append([match])

                    if regex_name in summary_dict:
                        if match in summary_dict[regex_name]:
                            summary_dict[regex_name][match]['count'] += 1
                            summary_dict[regex_name][match]['packets'].append(packet_counter)
                        else:
                            summary_dict[regex_name][match] = {'count': 1, 'packets': [packet_counter]}
                    else:
                        summary_dict[regex_name] = {match: {'count': 1, 'packets': [packet_counter]}}

        packet_counter += 1

    for regex_name, matches in summary_dict.items():
        for match, data in matches.items():
            packet_numbers = ', '.join(map(str, data['packets']))
            summary_sheet.append([regex_name, match, data['count'], packet_numbers])

    return workbook

def calculate_file_hash(file_path):
    """
    Calculate the SHA256 hash of a file.

    Parameters:
    - file_path (str): The path to the file.

    Returns:
    - str: The SHA256 hash of the file.

    Example:
    >>> calculate_file_hash('path/to/file.txt')
    '3a5b8c7d9e1f2a4b6c8d0e2f4a6b8c0d2e4f6a8b0c2d4e6f8a0b2c4d6e8f0a'

    Note:
    - This function reads the file in chunks of 8192 bytes to minimize memory usage.
    - The file is read in binary mode ('rb') to ensure compatibility with all file types.
    """
    hasher = hashlib.sha256()
    with open(file_path, 'rb') as file:
        while chunk := file.read(8192):
            hasher.update(chunk)
    file_hash = hasher.hexdigest()
    return file_hash

file_path = ''
output_file = file_path[:-5]+'_output.xlsx'
cap = pyshark.FileCapture(file_path)
regex_dict = initialize_regex()
workbook = Workbook()
workbook = process_packets(cap, regex_dict, workbook)
workbook.remove(workbook['Sheet'])
workbook.save(output_file)
cap.close()
